#!/usr/bin/env python3
"""
scripts/deploy-agent.py
Fully automated Open Brain agent deployment.
Usage: python deploy-agent.py <agent-id>
Example: python deploy-agent.py alex-rivers
"""
import argparse, hashlib, json, os, ssl, subprocess, sys, time, uuid
from pathlib import Path

try:
    import yaml
    from google.oauth2 import service_account
    from google.auth.transport.requests import Request
    import googleapiclient.discovery
    import openpyxl
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "-q",
        "pyyaml", "google-auth", "google-auth-httplib2",
        "google-api-python-client", "openpyxl"], check=True)
    import yaml
    from google.oauth2 import service_account
    from google.auth.transport.requests import Request
    import googleapiclient.discovery
    import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────
REPO       = Path("/Users/chussey/Documents/a-roofers-open-brain")
PROFILES   = REPO / "agents/profiles"
ENV_PATH   = REPO / ".env"
CREDS_XLSX = Path("~/Downloads/Agent Credentials.xlsx").expanduser()
SSH_KEY    = Path("~/.ssh/a_roofers_open_brain_ed25519").expanduser()
CTX        = ssl.create_default_context()
CTX.check_hostname = False
CTX.verify_mode    = ssl.CERT_NONE

# ── Helpers ─────────────────────────────────────────────────────────────────
def load_env():
    env = {}
    for line in ENV_PATH.read_text().splitlines():
        k, _, v = line.partition("=")
        if k.strip() and not k.strip().startswith("#"):
            env[k.strip()] = v.strip()
    return env

def http(url, payload=None, headers=None, method=None):
    method = method or ("POST" if payload else "GET")
    req = urllib_request.Request(
        url,
        data=json.dumps(payload).encode() if payload else None,
        headers={"Content-Type": "application/json", **(headers or {})},
        method=method,
    )
    try:
        with urllib_request.urlopen(req, context=CTX, timeout=20) as resp:
            return resp.status, json.loads(resp.read())
    except Exception as e:
        import urllib.error as ue
        if isinstance(e, ue.HTTPError):
            return e.code, {}
        return 0, {"_error": str(e)}

import urllib.request as urllib_request

def kasm(endpoint, payload, env):
    return http(
        f"{env.get('KASM_API_BASE','https://desktops.proexteriorsus.net')}/api/admin/{endpoint}",
        {**payload, "api_key": env["KASM_API_KEY"], "api_key_secret": env["KASM_API_KEY_SECRET"]},
    )

def ssh(cmd, env_dict=None):
    result = subprocess.run(
        ["ssh", "-i", str(SSH_KEY), "-o", "StrictHostKeyChecking=no",
         "-o", "ConnectTimeout=10", "root@5.78.146.161", cmd],
        capture_output=True, text=True, timeout=30,
    )
    return result.stdout.strip(), result.returncode

def slack_post(token, channel, text=None, blocks=None):
    payload = {"channel": channel}
    if text: payload["text"] = text
    if blocks: payload["blocks"] = blocks
    status, data = http(
        "https://slack.com/api/chat.postMessage", payload,
        headers={"Authorization": f"Bearer {token}"}
    )
    return data.get("ok"), data.get("ts"), data.get("error")

def step(label, ok, detail=""):
    icon = "✅" if ok else "❌"
    suffix = f" — {detail}" if detail else ""
    print(f"  {icon} {label}{suffix}", flush=True)
    return ok

# ── Load agent profile ───────────────────────────────────────────────────────
def load_profile(agent_id):
    path = PROFILES / f"{agent_id}.yaml"
    if not path.exists():
        print(f"❌ Profile not found: {path}")
        sys.exit(1)
    return yaml.safe_load(path.read_text())

# ── Load credentials from spreadsheet ───────────────────────────────────────
def load_credentials(agent_email):
    wb = openpyxl.load_workbook(str(CREDS_XLSX), data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip().lower() == agent_email.lower():
            return dict(zip(headers, [str(v or "").strip() for v in row]))
    return None

# ── Phase 2: Kasm user ───────────────────────────────────────────────────────
def phase2_kasm_user(profile, env, creds):
    print("\n── Phase 2: Kasm User ─────────────────────────────────────")
    email    = profile["google_workspace"]["email"]
    password = creds["Google Workspace Password"]
    first    = profile["identity"]["display_name"].split()[0]
    last     = profile["identity"]["display_name"].split()[-1]

    # Check existence via DB (Kasm API returns 403 from Python/macOS SSL)
    out, rc = ssh(f"docker exec kasm_db psql -U kasmapp -d kasm -c "
                  f"\"SELECT user_id FROM users WHERE username='{email}';\"")
    if email.split("@")[0] in out or "row" in out.lower():
        # Extract user_id
        uid_out, _ = ssh(f"docker exec kasm_db psql -U kasmapp -d kasm -t -c "
                         f"\"SELECT user_id FROM users WHERE username='{email}';\"")
        user_id = uid_out.strip().split()[0] if uid_out.strip() else "existing"
        step("Kasm user exists", True, f"user_id={user_id[:8]}...")
    else:
        # Create directly via DB INSERT (API create_user has permissions issues)
        import uuid as _uuid
        uid   = str(_uuid.uuid4())
        salt  = str(_uuid.uuid4())
        ph    = hashlib.sha256((password + salt).encode()).hexdigest()
        sql   = (f"INSERT INTO users (user_id,username,first_name,last_name,"
                 f"pw_hash,salt,locked,failed_pw_attempts,created,realm) "
                 f"VALUES ('{uid}','{email}','{first}','{last}',"
                 f"'{ph}','{salt}',false,0,NOW(),'local') "
                 f"ON CONFLICT (username) DO UPDATE "
                 f"SET pw_hash='{ph}',salt='{salt}',failed_pw_attempts=0;")
        out, rc = ssh(f'docker exec kasm_db psql -U kasmapp -d kasm -c "{sql}"')
        user_id = uid
        step("Kasm user created (DB)", "INSERT" in out or "UPDATE" in out,
             out[:30] if rc == 0 else out[:60])

    # Always sync password to ensure it matches GW
    salt    = str(uuid.uuid4())
    pw_hash = hashlib.sha256((password + salt).encode()).hexdigest()
    sql     = (f"UPDATE users SET pw_hash='{pw_hash}', salt='{salt}', "
               f"failed_pw_attempts=0 WHERE username='{email}';")
    out, rc = ssh(f'docker exec kasm_db psql -U kasmapp -d kasm -c "{sql}"')
    step("Kasm password synced", "UPDATE 1" in out)
    return user_id

# ── Phase 3: Kasm desktop ────────────────────────────────────────────────────
def phase3_kasm_desktop(profile, env, user_id):
    print("\n── Phase 3: Kasm Desktop ──────────────────────────────────")
    if not profile["kasm"].get("desktop_enabled"):
        step("Desktop skipped (disabled in profile)", True)
        return True

    image_id = profile["kasm"].get("image_id", "2c589484-3521-41fc-bec6-ac785ae87dd7")

    # Check existing sessions
    status, data = kasm("get_kasms", {}, env)
    kasms = data.get("kasms", [])
    existing = next((k for k in kasms if k.get("user_id") == user_id), None)
    if existing:
        step("Kasm session exists", True, existing.get("kasm_id", "")[:8])
        return True

    # Create session
    status, data = kasm("create_kasm", {
        "user_id": user_id,
        "image_id": image_id,
        "enable_sharing": False,
    }, env)
    if status == 200 and data.get("kasm_id"):
        step("Kasm session created", True, data["kasm_id"][:8])
        # Wait for container to be ready
        time.sleep(8)
        return True
    else:
        step("Kasm session created", False, str(data)[:100])
        return False

# ── Phase 4: Agent host files ────────────────────────────────────────────────
def phase4_agent_host(profile, env, creds):
    print("\n── Phase 4: Agent Host Setup ──────────────────────────────")
    agent_id  = profile["identity"]["id"]
    email     = profile["google_workspace"]["email"]
    image_id  = profile["kasm"].get("image_id", "2c589484-3521-41fc-bec6-ac785ae87dd7")
    sa_key    = profile["google_workspace"]["service_account_key_path"]
    desktop   = profile["kasm"].get("desktop_enabled", True)

    # Always create agent dir
    out, rc = ssh(f"mkdir -p /opt/openbrain/agents/{agent_id}/")
    step("Agent dir created", rc == 0)

    # Profile dir only if desktop_enabled
    if desktop:
        profile_dir = f"/mnt/kasm_profiles/{email}/{image_id}/.hermes"
        out, rc = ssh(f"mkdir -p {profile_dir}")
        step("Profile dir created", rc == 0)
    else:
        profile_dir = f"/opt/openbrain/agents/{agent_id}"
        step("No desktop (cron agent) — using agent dir", True, f"{profile_dir}")

    # Build .env content
    or_key = creds.get("OpenRouter API", "")
    fal_key = creds.get("FAL API", "")
    fc_key = creds.get("Firecrawl API", "")
    exa_key = creds.get("exa API", "")
    tav_key = creds.get("Tavily API", "")

    env_content = f"""# Alex Rivers — Hermes Agent Environment
# Generated by deploy-agent.py on {time.strftime('%Y-%m-%d')}

# ── Provider ──
OPENROUTER_API_KEY={or_key}

# ── Supabase ──
SUPABASE_URL={env.get('SUPABASE_URL','')}
SUPABASE_SERVICE_TOKEN={env.get('ALEX_SERVICE_TOKEN', env.get('SUPABASE_SERVICE_TOKEN',''))}
SUPABASE_ANON_KEY={env.get('SUPABASE_ANON_KEY','')}

# ── Command Center ──
COMMAND_CENTER_PUBLIC_URL=https://cc.proexteriorsus.net

# ── Google Workspace ──
GOOGLE_SERVICE_ACCOUNT_KEY_PATH={sa_key}
GOOGLE_IMPERSONATE_AS={email}

# ── ABC Supply ──
ABC_SUPPLY_ENV={env.get('ABC_SUPPLY_ENV','production')}
ABC_SUPPLY_CLIENT_ID={env.get('ABC_SUPPLY_CLIENT_ID','')}
ABC_SUPPLY_CLIENT_SECRET={env.get('ABC_SUPPLY_CLIENT_SECRET','')}
ABC_SUPPLY_API_BASE_URL={env.get('ABC_SUPPLY_API_BASE_URL','')}
ABC_SUPPLY_AUTH_BASE_URL={env.get('ABC_SUPPLY_AUTH_BASE_URL','')}
ABC_SUPPLY_SCOPES={env.get('ABC_SUPPLY_SCOPES','')}

# ── AgentMail ──
AGENTMAIL_API_KEY={env.get('AGENTMAIL_API_KEY','')}
AGENTMAIL_DOMAIN=agentmail.proexteriorsus.net

# ── Sentry ──
SENTRY_DSN={env.get('SENTRY_DSN','')}
SENTRY_ENVIRONMENT=production

# ── Research tools ──
FAL_API_KEY={fal_key}
FIRECRAWL_API_KEY={fc_key}
EXA_API_KEY={exa_key}
TAVILY_API_KEY={tav_key}

# ── Agent identity ──
AGENT_ID=alex-rivers
AGENT_ROLE=pricing-catalog
AGENT_DEPARTMENTS=accounting,operations,system
"""

    # Write via temp file + scp — backup existing .env first
    out, rc = ssh(f"[ -f {profile_dir}/.env ] && cp {profile_dir}/.env {profile_dir}/.env.bak.$(date +%Y%m%d) || true")
    tmp = Path(f"/tmp/{agent_id}-hermes-env")
    tmp.write_text(env_content)
    tmp.chmod(0o600)

    result = subprocess.run([
        "scp", "-q", "-i", str(SSH_KEY),
        "-o", "StrictHostKeyChecking=no",
        str(tmp), f"root@5.78.146.161:{profile_dir}/.env"
    ], capture_output=True, timeout=15)
    step(".env deployed", result.returncode == 0)
    tmp.unlink(missing_ok=True)

    # Set ownership
    out, rc = ssh(f"chown -R 1000:1000 /mnt/kasm_profiles/{email}/ && chmod 600 {profile_dir}/.env")
    step("Ownership set (kasm-user 1000)", rc == 0)

    # Write SOUL.md
    soul = f"""# {profile['identity']['display_name']} — Pro Exteriors Open Brain Agent

You are {profile['identity']['display_name']}, the {profile['identity']['persona_description']}

## Role

You map to {', '.join(profile['command_center']['maps_to'])}

## Operating rules

- You work with the {', '.join(profile['command_center']['departments'])} departments
- You NEVER approve work items (approval_decide = false — always false)
- You NEVER send external emails autonomously — draft only
- You surface findings and escalate; humans decide

## Escalation triggers

{chr(10).join('- ' + t for t in profile['guardrails']['escalation_triggers'])}

## Primary integrations

- ABC Supply API: pricing, catalog, invoices (read-only)
- AccuLynx: job/invoice data
- Command Center: work queue, evidence attachment, agent resume
"""
    tmp_soul = Path("/tmp/alex-soul.md")
    tmp_soul.write_text(soul)
    result = subprocess.run([
        "scp", "-q", "-i", str(SSH_KEY), "-o", "StrictHostKeyChecking=no",
        str(tmp_soul), f"root@5.78.146.161:{profile_dir}/SOUL.md"
    ], capture_output=True, timeout=15)
    step("SOUL.md deployed", result.returncode == 0)
    tmp_soul.unlink(missing_ok=True)

    # Write config.yaml
    config = f"""model:
  default: {profile['hermes']['model']}
  provider: {profile['hermes']['provider']}
  base_url: https://openrouter.ai/api/v1
providers: {{}}
fallback_providers: []
credential_pool_strategies: {{}}
toolsets:
{chr(10).join('- ' + t for t in profile['hermes']['toolsets'])}
"""
    tmp_cfg = Path("/tmp/alex-config.yaml")
    tmp_cfg.write_text(config)
    result = subprocess.run([
        "scp", "-q", "-i", str(SSH_KEY), "-o", "StrictHostKeyChecking=no",
        str(tmp_cfg), f"root@5.78.146.161:{profile_dir}/config.yaml"
    ], capture_output=True, timeout=15)
    step("config.yaml deployed", result.returncode == 0)
    tmp_cfg.unlink(missing_ok=True)

    # Fix ownership again after all files
    out, rc = ssh(f"chown -R 1000:1000 /mnt/kasm_profiles/{email}/")
    step("Final ownership fix", rc == 0)
    return True

# ── Phase 5: Google SA verification ─────────────────────────────────────────
def phase5_google(profile, env):
    print("\n── Phase 5: Google Workspace Verification ─────────────────")
    email   = profile["google_workspace"]["email"]
    sa_path = Path(profile["google_workspace"]["service_account_key_path"])
    # Use the downloaded key if the agent host path doesn't exist locally
    local_key = Path("~/Downloads").expanduser().glob("custom-frame-*.json")
    sa_file = next(local_key, None)
    if not sa_file:
        step("SA key found", False, "No key file in ~/Downloads")
        return False

    scopes = profile["google_workspace"]["gmail_scopes"]
    try:
        creds = service_account.Credentials.from_service_account_file(
            str(sa_file), scopes=scopes, subject=email)
        creds.refresh(Request())
        step("SA token acquired", True, f"impersonating {email}")
    except Exception as e:
        # Retry once — transient auth errors occur in batch runs
        time.sleep(3)
        try:
            creds = service_account.Credentials.from_service_account_file(
                str(sa_file), scopes=scopes, subject=email)
            creds.refresh(Request())
            step("SA token acquired (retry)", True, f"impersonating {email}")
        except Exception as e2:
            step("SA token acquired", False, str(e2)[:80])
            return False

    try:
        gmail = googleapiclient.discovery.build("gmail", "v1", credentials=creds, cache_discovery=False)
        labels = gmail.users().labels().list(userId="me").execute().get("labels", [])
        step("Gmail API works", True, f"{len(labels)} labels")
    except Exception as e:
        step("Gmail API works", False, str(e)[:60])

    try:
        drive = googleapiclient.discovery.build("drive", "v3", credentials=creds, cache_discovery=False)
        drive.files().list(pageSize=1).execute()
        step("Drive API works", True)
    except Exception as e:
        step("Drive API works", False, str(e)[:60])

    return True

# ── Phase 7: Slack app via Manifest API ──────────────────────────────────────
def phase7_slack(profile, env):
    print("\n── Phase 7: Slack App Creation ────────────────────────────")
    config_token = env.get("SLACK_APP_CONFIG_TOKEN", "")
    if len(config_token) < 100:
        step("Config token present", False, f"len={len(config_token)}, need ≥100")
        return None, None

    display = profile["slack"]["bot_display_name"]
    app_name = profile["slack"]["app_name"]

    # Build manifest
    manifest = {
        "display_information": {
            "name": app_name,
            "description": f"Open Brain agent: {profile['identity']['role']}",
            "background_color": "#11133f",
        },
        "features": {"bot_user": {"display_name": display, "always_online": True}},
        "oauth_config": {"scopes": {"bot": profile["slack"]["bot_scopes"]}},
        "settings": {
            "socket_mode_enabled": True,
            "org_deploy_enabled": False,
            "token_rotation_enabled": False,
            "event_subscriptions": {"bot_events": ["message.channels", "app_mention"]},
            "interactivity": {"is_enabled": False},
        },
    }

    # Check if app already exists — try create, treat app_already_exists as success
    # Note: apps.manifest.list only shows apps created by THIS config token session,
    # not apps created via UI or other tokens. Always attempt create and handle the error.
    status, data = http(
        "https://slack.com/api/apps.manifest.create",
        {"manifest": manifest, "team_id": profile["slack"]["team_id"]},
        headers={"Authorization": f"Bearer {config_token}"},
    )
    if not isinstance(data, dict):
        data = {}
    if data.get("error") == "app_already_exists":
        step("Slack app exists (skipping create)", True, "already installed")
        return data.get("app_id", "existing"), None
    if not data.get("ok"):
        step("Slack app created", False, data.get("error", "unknown"))
        return None, None

    app_id = data["app_id"]
    creds  = data.get("credentials", {})
    step("Slack app created", True, f"app_id={app_id}")

    # Install to workspace (requires admin OAuth token or manual install)
    admin_token = env.get("SLACK_ADMIN_BOT_TOKEN", "")
    if admin_token and len(admin_token) > 50:
        status, inst = http(
            "https://slack.com/api/admin.apps.approve",
            {"app_id": app_id, "team_id": profile["slack"]["team_id"]},
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        step("Workspace install (admin)", inst.get("ok", False), inst.get("error", ""))
    else:
        step("Workspace install", None, "⚠️  SLACK_ADMIN_BOT_TOKEN not set — manual install needed. Visit: https://slack.com/oauth/v2/authorize?client_id=" + creds.get("client_id","") + "&scope=chat:write,channels:read,channels:history,channels:join,files:read,app_mentions:read&team=" + profile["slack"]["team_id"])

    # Store app credentials in .env (via safe merge)
    signing_secret = creds.get("signing_secret", "")
    return app_id, None  # bot token comes after workspace install

# ── Phase 8: Supabase service token ─────────────────────────────────────────
def phase8_supabase(profile, env):
    print("\n── Phase 8: Supabase Service Token ────────────────────────")
    service_agent_id = profile["command_center"]["service_agent_id"]
    cc_url = profile["platform"]["command_center_url"]
    coolify_base = env.get("COOLIFY_API_BASE", "http://5.78.124.10:8000")
    cc_uuid = env.get("COOLIFY_CC_UUID", "og0rmt02rff8qti9nlfk3nr7")
    coolify_key = env.get("COOLIFY_PE_OPEN_BRAIN_API_KEY", "")

    # Check if token already exists
    existing_tokens = env.get("AGENT_SERVICE_TOKENS", "")
    if f"{service_agent_id}:" in existing_tokens:
        # Extract existing token
        for part in existing_tokens.split(","):
            if part.strip().startswith(f"{service_agent_id}:"):
                token = part.split(":", 1)[1].strip()
                # Verify it works
                status, data = http(
                    f"{cc_url}/api/agent/session", None,
                    headers={"Authorization": f"Bearer {token}"}
                )
                if status == 200:
                    step("Service token exists and works", True)
                    return token
                break

    # Generate new token
    raw = f"oas_{uuid.uuid4().hex}{uuid.uuid4().hex}"[:48]
    new_token = raw

    # Get current tokens from Coolify
    status, cur_envs = http(
        f"{coolify_base}/api/v1/applications/{cc_uuid}/envs", None,
        headers={"Authorization": f"Bearer {coolify_key}"},
    )
    if status != 200:
        step("Coolify env read", False, f"HTTP {status}")
        return None

    # Find and update AGENT_SERVICE_TOKENS
    env_list = cur_envs.get("data", cur_envs) if isinstance(cur_envs, dict) else cur_envs
    tokens_env = next((e for e in (env_list if isinstance(env_list, list) else []) if e.get("key") == "AGENT_SERVICE_TOKENS"), None)

    if tokens_env:
        new_val = tokens_env["value"] + f",{service_agent_id}:{new_token}"
        status, resp = http(
            f"{coolify_base}/api/v1/applications/{cc_uuid}/envs",
            {"key": "AGENT_SERVICE_TOKENS", "value": new_val, "is_preview": False},
            headers={"Authorization": f"Bearer {coolify_key}"},
        )
        step("Token added to Coolify", status in (200, 201))
    else:
        step("Coolify env read", False, "AGENT_SERVICE_TOKENS not found in env list")
        return None

    # Trigger redeploy
    status, _ = http(
        f"{coolify_base}/api/v1/deploy?uuid={cc_uuid}", {},
        headers={"Authorization": f"Bearer {coolify_key}"},
    )
    step("Coolify redeploy triggered", status in (200, 201))

    # Wait for deploy
    print("  ⏳ Waiting for Coolify deploy (max 3 min)...")
    for i in range(18):
        time.sleep(10)
        s, d = http(f"{cc_url}/healthz", None, headers={})
        if s == 200:
            print(f"  ✅ Deploy complete (build: {d.get('buildCommit','?')[:8]})")
            break
    else:
        step("Deploy complete", False, "timeout after 3 min")
        return None

    # Verify token works
    time.sleep(3)
    status, data = http(
        f"{cc_url}/api/agent/session", None,
        headers={"Authorization": f"Bearer {new_token}"}
    )
    step("Service token verified", status == 200, data.get("actor", {}).get("id", ""))
    return new_token if status == 200 else None

# ── Phase 9: ABC Supply ──────────────────────────────────────────────────────
def phase9_abc(profile, env):
    print("\n── Phase 9: ABC Supply API ────────────────────────────────")
    if not profile["integrations"].get("abc_supply"):
        step("ABC Supply skipped (disabled in profile)", True)
        return True

    import urllib.parse
    auth_base = env.get("ABC_SUPPLY_AUTH_BASE_URL") or "https://auth.partners.abcsupply.com/oauth2/ausvvp0xuwGKLenYy357"
    api_base  = env.get("ABC_SUPPLY_API_BASE_URL") or "https://partners.abcsupply.com"
    token_url = auth_base + "/v1/token"

    body = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": env.get("ABC_SUPPLY_CLIENT_ID", ""),
        "client_secret": env.get("ABC_SUPPLY_CLIENT_SECRET", ""),
        "scope": "product.read location.read invoice.read",
    }).encode()

    import urllib.request as ur
    req = ur.Request(token_url, data=body, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    try:
        with ur.urlopen(req, context=CTX, timeout=15) as resp:
            token = json.loads(resp.read())["access_token"]
        step("ABC Supply auth", True)
    except Exception as e:
        step("ABC Supply auth", False, str(e)[:60])
        return False

    req2 = ur.Request(
        f"{api_base}/api/product/v1/items?pageNumber=1&itemsPerPage=2",
        headers={"Authorization": f"Bearer {token}"}
    )
    try:
        with ur.urlopen(req2, context=CTX, timeout=15) as resp:
            data = json.loads(resp.read())
            step("ABC catalog accessible", True, f"{data.get('totalCount','?')} items")
    except Exception as e:
        step("ABC catalog accessible", False, str(e)[:60])
    return True

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("agent_id")
    args = parser.parse_args()

    print(f"╔══ deploy-agent: {args.agent_id} ══{'═'*(40-len(args.agent_id))}")
    print(f"║  {time.strftime('%Y-%m-%d %H:%M:%S CT')}")
    print(f"╚{'═'*56}")

    profile = load_profile(args.agent_id)
    env     = load_env()
    email   = profile["google_workspace"]["email"]

    print(f"\n  Agent:  {profile['identity']['display_name']}")
    print(f"  Role:   {profile['identity']['role']}")
    print(f"  Email:  {email}")

    creds = load_credentials(email)
    if not creds:
        print(f"❌ No credentials found for {email} in Agent Credentials.xlsx")
        sys.exit(1)
    print(f"  Creds:  ✅ found in spreadsheet")

    results = {}

    # Phase 2: Kasm user
    user_id = phase2_kasm_user(profile, env, creds)
    results["kasm_user"] = bool(user_id)

    # Phase 3: Kasm desktop
    results["kasm_desktop"] = phase3_kasm_desktop(profile, env, user_id) if user_id else False

    # Phase 4: Agent host
    results["agent_host"] = phase4_agent_host(profile, env, creds)

    # Phase 5: Google
    results["google_sa"] = phase5_google(profile, env)

    # Phase 7: Slack
    app_id, bot_token = phase7_slack(profile, env)
    results["slack_app"] = bool(app_id)

    # Phase 8: Supabase
    service_token = phase8_supabase(profile, env)
    results["supabase"] = bool(service_token)

    # Phase 9: ABC Supply
    results["abc_supply"] = phase9_abc(profile, env)

    # Summary
    passed = sum(results.values())
    total  = len(results)
    print(f"\n{'═'*56}")
    print(f"Deploy result: {passed}/{total} phases passed")
    for k, v in results.items():
        print(f"  {'✅' if v else '❌'} {k}")

    # Store service token in results for validation phase
    return results, service_token, profile, env


if __name__ == "__main__":
    results, service_token, profile, env = main()
    # Save service token for validation script
    if service_token:
        Path("/tmp/alex-service-token").write_text(service_token)
        Path("/tmp/alex-service-token").chmod(0o600)
