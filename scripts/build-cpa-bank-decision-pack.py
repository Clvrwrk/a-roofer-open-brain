#!/usr/bin/env python3
"""
CPA Bank Decision Pack — overnight / pre-CPA meeting.

Purpose (Chris + CPA, next morning):
  Decide what GOES / HOLDS / NEVER GOES to the LOC bank for the
  first full cashflow review (target: 2026-08-05).

Not a final bank package. Decision mock with:
  - Live cash-basis facts from QBO mirror + AccuLynx WIP pack
  - Explicit INCLUDE/EXCLUDE/HOLD recommendations
  - Placeholder mock-accrual tabs (policy TBD with CPA)
  - Recon flags (QB AR vs AccuLynx open AR, etc.)

Hard rules:
  - Read-only on QBO (mirror + read-only client only)
  - Never print secrets
  - Every money tab states basis: CASH | MOCK_ACCRUAL_TBD | RECON
"""

from __future__ import annotations

import json
import os
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
OUT_DIR = REPO / "outputs" / "cpa-bank-decision-pack"
WIP_DIR = REPO / "outputs" / "chandler-wip-v1"
PAGE = 1000

# Decision codes for CPA walkthrough
SEND = "SEND"
HOLD = "HOLD"
NEVER = "NEVER"
NEEDS_CPA = "NEEDS_CPA"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SEND_FILL = PatternFill("solid", fgColor="C6EFCE")
HOLD_FILL = PatternFill("solid", fgColor="FFEB9C")
NEVER_FILL = PatternFill("solid", fgColor="FFC7CE")
NEEDS_FILL = PatternFill("solid", fgColor="BDD7EE")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
TITLE_FONT = Font(bold=True, size=16)
SECTION_FONT = Font(bold=True, size=12)
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _load_env_file(path: Path, *, override: bool = False) -> None:
    if not path.exists():
        return
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        v = v.strip().strip('"').strip("'")
        if not k:
            continue
        if override or k not in os.environ:
            os.environ[k] = v


def load_env() -> None:
    _load_env_file(Path.home() / ".config/cleverwork/master.env", override=False)
    _load_env_file(REPO / ".env", override=True)
    _load_env_file(REPO / "config" / ".env", override=True)


def require_env(*names: str) -> str:
    for name in names:
        val = os.environ.get(name)
        if val:
            return val
    raise SystemExit(f"Missing env: one of {names}")


def sb_headers(key: str) -> Dict[str, str]:
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Accept": "application/json",
        "Prefer": "count=exact",
    }


def sb_get(base: str, key: str, path_and_qs: str) -> Any:
    url = f"{base.rstrip('/')}/rest/v1/{path_and_qs}"
    req = urllib.request.Request(url, headers=sb_headers(key))
    with urllib.request.urlopen(req, timeout=120) as resp:
        return json.loads(resp.read().decode())


def sb_get_all(base: str, key: str, table: str, select: str, order: Optional[str] = None) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    start = 0
    while True:
        params: Dict[str, str] = {"select": select}
        if order:
            params["order"] = order
        qs = urllib.parse.urlencode(params, safe="(),.*:")
        url = f"{base.rstrip('/')}/rest/v1/{table}?{qs}"
        req = urllib.request.Request(
            url,
            headers={**sb_headers(key), "Range": f"{start}-{start + PAGE - 1}", "Range-Unit": "items"},
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            chunk = json.loads(resp.read().decode())
        if not isinstance(chunk, list):
            raise SystemExit(f"Unexpected response for {table}")
        rows.extend(chunk)
        if len(chunk) < PAGE:
            break
        start += PAGE
        if start > 50_000:
            break
    return rows


def num(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def money(v: Any) -> float:
    return round(num(v), 2)


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 12
        for cell in col[:80]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def style_header_row(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def paint_decision(cell) -> None:
    v = (cell.value or "").upper()
    if v == SEND:
        cell.fill = SEND_FILL
    elif v == HOLD:
        cell.fill = HOLD_FILL
    elif v == NEVER:
        cell.fill = NEVER_FILL
    elif v == NEEDS_CPA:
        cell.fill = NEEDS_FILL


def write_table(ws, headers: List[str], rows: List[List[Any]], start_row: int = 1) -> int:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    r = start_row + 1
    for row in rows:
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.border = THIN
            if headers[c - 1].lower() in ("decision", "cpa decision", "bank decision"):
                paint_decision(cell)
        r += 1
    autosize(ws)
    return r


def latest_wip_path() -> Path:
    files = sorted(WIP_DIR.glob("chandler-wip-*.xlsx"), reverse=True)
    if not files:
        raise SystemExit(f"No WIP pack in {WIP_DIR} — run build-chandler-wip-v1.py first")
    return files[0]


def load_wip_summary(path: Path) -> Dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    stages = []
    for row in wb["By Stage"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            stages.append({"stage": row[0], "jobs": row[1] or 0, "contract": num(row[2]), "ar": num(row[3])})
    locs = []
    for row in wb["By Location"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            locs.append({"location": row[0], "jobs": row[1] or 0, "contract": num(row[2]), "ar": num(row[3])})
    reps = []
    for row in wb["By Salesperson"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            reps.append({"rep": row[0], "jobs": row[1] or 0, "contract": num(row[2]), "ar": num(row[3])})
    cash_accounts = []
    cash_ws = wb["Cash Snapshot (QB)"]
    # After header Account/Amount
    for row in cash_ws.iter_rows(min_row=1, values_only=True):
        if row and row[0] == "Account":
            continue
        if row and isinstance(row[0], str) and row[1] is not None and row[0] not in (
            "Company", "Legal name", "Realm", "API base", "Read-only enforced",
            "Cash-like total", "Notes", "Cash/Bank account lines (sample)",
            "QuickBooks cash snapshot (read-only)", "Errors",
        ):
            if row[0] and not str(row[0]).startswith("Cash"):
                cash_accounts.append({"name": row[0], "amount": num(row[1])})
    cash_total = None
    company = None
    for row in cash_ws.iter_rows(min_row=1, max_row=15, values_only=True):
        if row and row[0] == "Cash-like total":
            cash_total = num(row[1])
        if row and row[0] == "Company":
            company = row[1]
    exceptions = []
    for row in wb["Exception Queue"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            exceptions.append(list(row))
    jobs = list(wb["Jobs"].iter_rows(min_row=2, values_only=True))
    wb.close()
    open_ar = sum(num(r[12]) for r in jobs)
    active = [r for r in jobs if (r[4] or "").lower() != "closed"]
    return {
        "path": str(path),
        "stages": stages,
        "locations": locs,
        "reps": reps,
        "cash_accounts": cash_accounts,
        "cash_total": cash_total,
        "company": company,
        "exceptions": exceptions,
        "job_count": len(jobs),
        "active_job_count": len(active),
        "open_ar": open_ar,
        "active_contract": sum(num(r[10]) or num(r[11]) for r in active),
    }


def fetch_mirror(base: str, key: str) -> Dict[str, Any]:
    registers = sb_get_all(
        base, key, "v_qbo_register_accounts",
        "account_name,register_kind,current_balance",
        order="register_kind.asc,account_name.asc",
    )
    accounts = sb_get_all(
        base, key, "qbo_accounts",
        "name,account_type,account_sub_type,current_balance,active,classification",
        order="account_type.asc,name.asc",
    )
    # latest BS snapshot meta
    snaps = sb_get(
        base + "/",
        key,
        "qbo_report_snapshots?select=report_name,as_of_date,accounting_method,fetched_at&order=fetched_at.desc&limit=20",
    ) if False else None
    # use full path helper
    try:
        snaps = sb_get_all(base, key, "qbo_report_snapshots", "report_name,as_of_date,accounting_method,fetched_at")
    except Exception as e:
        snaps = []
        print("  report_snapshots:", e)

    # AR account
    ar_bal = None
    ap_bal = None
    for a in accounts:
        n = (a.get("name") or "").lower()
        if "accounts receivable" in n:
            ar_bal = num(a.get("current_balance"))
        if "accounts payable" in n and "a/p" in n:
            ap_bal = num(a.get("current_balance"))

    by_kind: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in registers:
        by_kind[r.get("register_kind") or "other"].append(r)

    bank_total = sum(num(r.get("current_balance")) for r in by_kind.get("bank", []))
    cc_total = sum(num(r.get("current_balance")) for r in by_kind.get("credit_card", []))
    np_total = sum(num(r.get("current_balance")) for r in by_kind.get("notes_payable", []))
    st_total = sum(num(r.get("current_balance")) for r in by_kind.get("short_term_financing", []))

    return {
        "registers": registers,
        "by_kind": by_kind,
        "accounts": accounts,
        "snaps": snaps or [],
        "qb_ar": ar_bal,
        "qb_ap": ap_bal,
        "bank_total": bank_total,
        "cc_total": cc_total,
        "np_total": np_total,
        "st_total": st_total,
    }


def build_workbook(wip: Dict[str, Any], mirror: Dict[str, Any]) -> Path:
    wb = Workbook()
    stamp = date.today().isoformat()
    gen = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    # ── 00 Cover ──────────────────────────────────────────────
    cover = wb.active
    cover.title = "00_Cover"
    cover["A1"] = "CPA · Bank Decision Pack (MOCK — NOT FOR BANK SEND)"
    cover["A1"].font = TITLE_FONT
    cover["A2"] = "Pro Exteriors LLC · Cash-basis books today · Mock accrual TBD with CPA"
    rows = [
        ("Generated", gen),
        ("Meeting", "CPA walkthrough — decide bank package contents"),
        ("Bank target", "2026-08-05 first full cashflow review (LOC)"),
        ("Also this week", "AR/WIP pack review tomorrow → Friday AR/WIP meeting"),
        ("WIP source", Path(wip["path"]).name),
        ("QB company", wip.get("company") or "Pro Exteriors LLC"),
        ("Books basis (live QBO)", "CASH"),
        ("Mock accrual", "NOT built yet — tabs are decision scaffolds only"),
        ("QBO access", "READ-ONLY forever (PEC-98)"),
        ("Linear trunk", "PEC-100 / CashFlow-WIP Executive Surface"),
    ]
    r = 4
    for k, v in rows:
        cover[f"A{r}"] = k
        cover[f"B{r}"] = v
        cover[f"A{r}"].font = Font(bold=True)
        r += 1
    r += 1
    cover[f"A{r}"] = "HOW TO USE THIS WORKBOOK (30 min with CPA)"
    cover[f"A{r}"].font = SECTION_FONT
    r += 1
    for line in [
        "1. Start on 01_Decision_Matrix — green SEND / yellow HOLD / red NEVER / blue NEEDS_CPA.",
        "2. Flip each SEND candidate tab; confirm numbers and footnotes make sense.",
        "3. On 02_Open_Questions — answer before Aug 5 bank pack is assembled.",
        "4. Mock accrual tabs (09–11) are intentionally thin: decide policy, not polish math yet.",
        "5. Nothing in this file is bank-ready until CPA decision column is filled and Chris signs Cover.",
    ]:
        cover[f"A{r}"] = line
        r += 1
    r += 1
    cover[f"A{r}"] = "HEADLINE FACTS (live, cash + WIP ops)"
    cover[f"A{r}"].font = SECTION_FONT
    r += 1
    facts = [
        ("Bank cash (register sum)", money(mirror["bank_total"])),
        ("Cash-like BS walk (WIP pack)", money(wip.get("cash_total"))),
        ("QB Accounts Receivable (CoA balance)", money(mirror.get("qb_ar"))),
        ("AccuLynx open AR (job invoices sum)", money(wip["open_ar"])),
        ("AR recon gap (AccuLynx − QB)", money(wip["open_ar"] - num(mirror.get("qb_ar")))),
        ("Credit cards total (register sum)", money(mirror["cc_total"])),
        ("Notes payable total (register sum)", money(mirror["np_total"])),
        ("Open jobs in WIP pack (excl cancelled/dead)", wip["job_count"]),
        ("Active (non-Closed milestone)", wip["active_job_count"]),
        ("Active contract $ (non-Closed)", money(wip["active_contract"])),
    ]
    for k, v in facts:
        cover[f"A{r}"] = k
        cover[f"B{r}"] = v
        if "gap" in k.lower():
            cover[f"A{r}"].fill = WARN_FILL
            cover[f"B{r}"].fill = WARN_FILL
        r += 1
    r += 1
    cover[f"A{r}"] = "SIGN-OFF (fill after CPA meeting)"
    cover[f"A{r}"].font = SECTION_FONT
    r += 1
    for label in ["Chris decision date", "CPA present", "Tabs approved for Aug 5 bank pack", "Tabs held for later", "Tabs never to bank", "Next build owner"]:
        cover[f"A{r}"] = label
        cover[f"B{r}"] = ""
        r += 1
    cover.column_dimensions["A"].width = 52
    cover.column_dimensions["B"].width = 56

    # ── 01 Decision Matrix ────────────────────────────────────
    dm = wb.create_sheet("01_Decision_Matrix")
    dm["A1"] = "Bank package decision matrix — CPA meeting"
    dm["A1"].font = TITLE_FONT
    dm["A2"] = "Edit column F (CPA Decision). Pre-fills in E are Chris/agent draft recommendations only."
    headers = [
        "Tab / Schedule",
        "Audience",
        "Basis",
        "What it answers",
        "Draft rec",
        "CPA Decision",
        "Why / notes",
        "Owner to finish by Aug 5",
    ]
    matrix = [
        ["00_Cover", "Bank+CPA", "META", "Scope, basis statement, sign-off", SEND, "", "Always include as cover letter inside pack", "Chris"],
        ["03_Cash_Position", "Bank", "CASH", "Where is cash and what debts sit on books?", SEND, "", "Bank wants liquidity picture", "Grok+Lucinda"],
        ["04_CF_Org_Template", "Bank", "CASH", "Org-level cashflow (direct method scaffold)", HOLD, "", "Need period selection + full register rollforward for Aug 5", "Grok"],
        ["05_CF_by_Location_TBD", "Bank", "CASH", "Location-level cashflow detail", HOLD, "", "Location SoT (AccuLynx account_key vs QBO Class) not locked", "Chris+CPA"],
        ["06_AR_Aging_QB", "Bank", "CASH", "QB A/R aging (books)", SEND, "", "Use books AR for bank; footnote AccuLynx ops AR", "Grok"],
        ["07_WIP_AR_Ops", "Bank+Ops", "OPS/CASH hybrid", "Job-stage WIP + AccuLynx open AR by location", NEEDS_CPA, "", "Powerful ops truth; may overstate vs cash books — CPA gate", "CPA+Chris"],
        ["08_Debt_LOC_NP", "Bank", "CASH", "LOC / notes / CC exposure", SEND, "", "Bank will ask; FUB notes + CCs material", "Grok+Lucinda"],
        ["09_Mock_Accrual_Policy", "CPA", "MOCK_TBD", "Recognition rules until Q1 2027 flip", NEEDS_CPA, "", "Policy only for tomorrow — not bank math yet", "CPA"],
        ["10_Mock_P&L_Scaffold", "Bank?", "MOCK_TBD", "Management earnings if bank wants accrual-like", HOLD, "", "Only if CPA says bank expects it Aug 5", "CPA"],
        ["11_Bridge_Cash_to_Mock", "CPA+Bank?", "RECON", "Cash earnings → mock accrual bridge", HOLD, "", "Required if mock P&L goes; else internal only", "CPA"],
        ["12_Data_Quality_Flags", "Internal", "META", "Gaps, match rates, recon warnings", SEND, "", "Bank may not see full tab — use as appendix or internal", "Grok"],
        ["13_Friday_AR_WIP_Agenda", "Internal", "OPS", "Friday meeting run-of-show from same spine", NEVER, "", "Ops only — not for bank", "Chris"],
        ["Full QBO P&L detail (all GL)", "Bank", "CASH", "Complete cash P&L for quarter", NEEDS_CPA, "", "May send standard QBO export instead of our sheet", "CPA"],
        ["Full QBO BS", "Bank", "CASH", "Balance sheet cash basis", SEND, "", "Standard BS + our cash/debt detail", "Lucinda/CPA"],
        ["Partner equity / distributions detail", "Bank", "CASH", "Owner draws", NEEDS_CPA, "", "Sensitive — CPA decides level of detail", "CPA"],
        ["Job-level customer names (PII-ish)", "Bank", "OPS", "Customer-level WIP", HOLD, "", "May aggregate or hash for bank; full names internal", "Chris"],
        ["ABC job-cost gap pack (~29% unmatched)", "Internal", "OPS", "Unposted ABC / register gaps", NEVER, "", "Ops cleanup — not bank narrative", "Alex/Lucinda"],
        ["Interactive CC /executive/cashflow-wip", "Internal", "OPS", "Product UI", NEVER, "", "Not ready; not a bank artifact", "Eng"],
    ]
    write_table(dm, headers, matrix, start_row=4)
    dm.row_dimensions[4].height = 30

    # ── 02 Open Questions ─────────────────────────────────────
    oq = wb.create_sheet("02_Open_Questions")
    oq["A1"] = "Open questions for CPA tomorrow (must answer for Aug 5)"
    oq["A1"].font = TITLE_FONT
    qrows = [
        ["Q1", "Does the bank require cash-basis only, or cash + management accrual view?", "", "Blocks mock accrual SEND"],
        ["Q2", "First reporting period for Aug 5 pack? (MTD Jul / Q2 / TTM / as-of date)", "", "Blocks CF tabs"],
        ["Q3", "Location dimension for bank: AccuLynx account_key, QBO Class, or legal entity only?", "", "Blocks 05_CF_by_Location"],
        ["Q4", "May we send AccuLynx WIP/AR schedule if we clearly label it as operational (not books)?", "", "Blocks 07"],
        ["Q5", "How to present AR recon: QB AR $ vs AccuLynx open AR $ (large gap)?", "", "Credibility risk"],
        ["Q6", "Level of debt detail: each FUB note vs rolled notes payable?", "", "08 tab"],
        ["Q7", "Partner distributions — full, summary, or omit?", "", "Sensitivity"],
        ["Q8", "Mock accrual revenue: % complete by stage, cost-to-cost, or milestone deposits?", "", "Policy for 2026 bridge"],
        ["Q9", "When CPA flips company to accrual Q1 2027, do we retire mock engine or map 1:1?", "", "Shelf-life design"],
        ["Q10", "Who signs bank pack: Chris only / Chris+CPA / Lucinda review?", "", "Process"],
    ]
    write_table(oq, ["#", "Question", "CPA answer", "If unanswered…"], qrows, start_row=3)

    # ── 03 Cash Position ──────────────────────────────────────
    cp = wb.create_sheet("03_Cash_Position")
    cp["A1"] = "Cash & liquidity position — CASH BASIS (QBO mirror registers)"
    cp["A1"].font = TITLE_FONT
    cp["A2"] = f"As-of mirror balances · Generated {gen} · NOT a full cashflow statement"
    cp["A3"] = "Decision draft: SEND to bank (with period note + as-of date locked before Aug 5)"
    cp["A3"].fill = SEND_FILL

    r = 5
    cp[f"A{r}"] = "BANKS"
    cp[f"A{r}"].font = SECTION_FONT
    r += 1
    bank_rows = [
        [x["account_name"], money(x.get("current_balance")), SEND, "Operating liquidity"]
        for x in mirror["by_kind"].get("bank", [])
    ]
    r = write_table(cp, ["Account", "Balance $", "Draft rec", "Note"], bank_rows, start_row=r)
    cp.cell(r, 1, "Total banks")
    cp.cell(r, 2, money(mirror["bank_total"]))
    cp.cell(r, 1).font = Font(bold=True)
    r += 2

    cp[f"A{r}"] = "CREDIT CARDS (liability balances as in CoA/register)"
    cp[f"A{r}"].font = SECTION_FONT
    r += 1
    cc_rows = []
    for x in sorted(mirror["by_kind"].get("credit_card", []), key=lambda z: abs(num(z.get("current_balance"))), reverse=True):
        bal = money(x.get("current_balance"))
        if bal == 0:
            rec = HOLD
            note = "Zero balance — optional detail"
        else:
            rec = SEND
            note = "Material facility / card exposure"
        cc_rows.append([x["account_name"], bal, rec, note])
    r = write_table(cp, ["Account", "Balance $", "Draft rec", "Note"], cc_rows, start_row=r)
    cp.cell(r, 1, "Total CC (register sum)")
    cp.cell(r, 2, money(mirror["cc_total"]))
    cp.cell(r, 1).font = Font(bold=True)
    r += 2

    cp[f"A{r}"] = "NOTES PAYABLE / LOANS"
    cp[f"A{r}"].font = SECTION_FONT
    r += 1
    np_rows = []
    for x in sorted(mirror["by_kind"].get("notes_payable", []), key=lambda z: abs(num(z.get("current_balance"))), reverse=True):
        bal = money(x.get("current_balance"))
        if bal == 0:
            rec, note = HOLD, "Zero — may hide retired notes"
        else:
            rec, note = SEND, "Include in debt schedule"
        np_rows.append([x["account_name"], bal, rec, note])
    r = write_table(cp, ["Account", "Balance $", "Draft rec", "Note"], np_rows, start_row=r)
    cp.cell(r, 1, "Total notes payable (register sum)")
    cp.cell(r, 2, money(mirror["np_total"]))
    r += 2
    cp[f"A{r}"] = "SHORT-TERM FINANCING"
    cp[f"A{r}"].font = SECTION_FONT
    r += 1
    st_rows = [
        [x["account_name"], money(x.get("current_balance")), SEND if num(x.get("current_balance")) else HOLD, ""]
        for x in mirror["by_kind"].get("short_term_financing", [])
    ]
    write_table(cp, ["Account", "Balance $", "Draft rec", "Note"], st_rows, start_row=r)

    # ── 04 CF Org Template ────────────────────────────────────
    cf = wb.create_sheet("04_CF_Org_Template")
    cf["A1"] = "Organization cashflow — TEMPLATE (direct method) · CASH BASIS"
    cf["A1"].font = TITLE_FONT
    cf["A2"] = "HOLD for Aug 5 until period locked. Structure bank asked for — numbers to fill from register rollforward."
    cf["A2"].fill = HOLD_FILL
    cf_headers = ["Line", "Section", "Jul MTD (TBD)", "Aug MTD (TBD)", "QTD (TBD)", "Source", "Draft rec"]
    cf_lines = [
        ["1", "Beginning cash", "", "", "", "Bank registers", SEND],
        ["2", "Customer collections / deposits", "", "", "", "QBO deposits + payments", SEND],
        ["3", "Other cash receipts", "", "", "", "Registers", SEND],
        ["4", "Total cash in", "", "", "", "Sum 2–3", SEND],
        ["5", "Materials (ABC + other suppliers)", "", "", "", "Registers / ABC", SEND],
        ["6", "Subcontractors / field labor cash", "", "", "", "Registers", SEND],
        ["7", "Payroll & payroll taxes", "", "", "", "Payroll system / QBO", SEND],
        ["8", "Operating expenses", "", "", "", "Registers", SEND],
        ["9", "Debt service (notes + LOC)", "", "", "", "Notes registers", SEND],
        ["10", "Owner distributions", "", "", "", "Equity accounts", NEEDS_CPA],
        ["11", "Total cash out", "", "", "", "Sum 5–10", SEND],
        ["12", "Net change in cash", "", "", "", "4 − 11", SEND],
        ["13", "Ending cash", "", "", "", "1 + 12; ties to 03", SEND],
        ["14", "LOC draws / (payments) detail", "", "", "", "LOC register", SEND],
        ["15", "Memo: non-cash WIP movement", "", "", "", "Ops only", NEVER],
    ]
    write_table(cf, cf_headers, cf_lines, start_row=4)
    cf["A22"] = "Build note: implement as SQL over qbo_registers lines once period dates locked with CPA."
    cf["A23"] = "Location version (tab 05) = same lines GROUP BY location SoT — blocked on Q3."

    # ── 05 Location TBD ───────────────────────────────────────
    loc = wb.create_sheet("05_CF_by_Location_TBD")
    loc["A1"] = "Cashflow by location — BLOCKED pending location SoT"
    loc["A1"].font = TITLE_FONT
    loc["A2"] = "Draft rec: HOLD. AccuLynx account_key rollups exist for WIP/AR; cash registers are mostly company-level."
    loc["A2"].fill = HOLD_FILL
    loc["A4"] = "What we CAN show today (ops, not cash CF)"
    loc["A4"].font = SECTION_FONT
    write_table(
        loc,
        ["Location (AccuLynx account_key)", "Jobs", "Contract/Est $", "Open AR $", "Bank CF available?"],
        [
            [x["location"], x["jobs"], money(x["contract"]), money(x["ar"]), "No — WIP/AR only"]
            for x in wip["locations"]
        ],
        start_row=5,
    )
    loc["A18"] = "Options for CPA:"
    loc["A19"] = "A) Bank gets org CF only until Class mapping exists"
    loc["A20"] = "B) Proxy location CF using job deposits/collections mapped via AccuLynx job # (approximate)"
    loc["A21"] = "C) Force QBO Class discipline starting Aug (forward-looking only)"

    # ── 06 AR Aging placeholder ───────────────────────────────
    ar = wb.create_sheet("06_AR_Aging_QB")
    ar["A1"] = "QB Accounts Receivable — CASH BOOKS"
    ar["A1"].font = TITLE_FONT
    ar["A2"] = f"CoA AR balance (mirror): {money(mirror.get('qb_ar'))} · Full aging detail = export AgedReceivables for Aug 5 pack"
    ar["A3"] = "Draft rec: SEND standard QBO AR aging PDF/CSV + one-line CoA balance here"
    ar["A3"].fill = SEND_FILL
    ar["A5"] = "CRITICAL RECON (do not ignore in CPA meeting)"
    ar["A5"].font = SECTION_FONT
    ar["A5"].fill = WARN_FILL
    write_table(
        ar,
        ["Metric", "Amount $", "Source", "Implication"],
        [
            ["QB A/R (CoA)", money(mirror.get("qb_ar")), "qbo_accounts", "Books receivable"],
            ["AccuLynx open AR (sum job invoices)", money(wip["open_ar"]), "WIP pack / acculynx_invoices", "Ops collection view"],
            ["Gap (AccuLynx − QB)", money(wip["open_ar"] - num(mirror.get("qb_ar"))), "Calc", "Do NOT send both without explanation"],
        ],
        start_row=6,
    )
    ar["A12"] = "CPA choices: (1) Bank sees only QB AR  (2) Bank sees QB AR + ops WIP appendix  (3) Reconcile before any send"
    ar["A13"] = "Recommendation: (1) for Aug 5 primary; (2) only if labeled OPERATIONAL and gap footnoted"

    # ── 07 WIP AR Ops ─────────────────────────────────────────
    wip_s = wb.create_sheet("07_WIP_AR_Ops")
    wip_s["A1"] = "Operational WIP & AR by stage/location — NOT cash books"
    wip_s["A1"].font = TITLE_FONT
    wip_s["A2"] = f"Source: {Path(wip['path']).name} · Stages provisional (docs/75 rename not applied in generator yet)"
    wip_s["A3"] = "Draft rec: NEEDS_CPA — powerful for Friday AR meeting; sensitive for bank"
    wip_s["A3"].fill = NEEDS_FILL
    r = 5
    wip_s[f"A{r}"] = "BY STAGE"
    wip_s[f"A{r}"].font = SECTION_FONT
    r += 1
    r = write_table(
        wip_s,
        ["Operating stage (v1 provisional)", "Jobs", "Contract/Est $", "Open AR $"],
        [[x["stage"], x["jobs"], money(x["contract"]), money(x["ar"])] for x in wip["stages"]],
        start_row=r,
    )
    r += 1
    wip_s[f"A{r}"] = "BY LOCATION"
    wip_s[f"A{r}"].font = SECTION_FONT
    r += 1
    r = write_table(
        wip_s,
        ["Location", "Jobs", "Contract/Est $", "Open AR $"],
        [[x["location"], x["jobs"], money(x["contract"]), money(x["ar"])] for x in wip["locations"]],
        start_row=r,
    )
    r += 1
    wip_s[f"A{r}"] = "TOP SALESPEOPLE BY CONTRACT"
    wip_s[f"A{r}"].font = SECTION_FONT
    r += 1
    write_table(
        wip_s,
        ["Salesperson", "Jobs", "Contract/Est $", "Open AR $"],
        [[x["rep"], x["jobs"], money(x["contract"]), money(x["ar"])] for x in wip["reps"][:15]],
        start_row=r,
    )

    # ── 08 Debt ───────────────────────────────────────────────
    debt = wb.create_sheet("08_Debt_LOC_NP")
    debt["A1"] = "Debt, cards, financing — CASH BASIS register balances"
    debt["A1"].font = TITLE_FONT
    debt["A2"] = "Draft rec: SEND rolled summary + material lines; CPA trims zero/noise accounts"
    debt["A2"].fill = SEND_FILL
    all_debt = []
    for kind in ("credit_card", "notes_payable", "short_term_financing"):
        for x in mirror["by_kind"].get(kind, []):
            bal = money(x.get("current_balance"))
            all_debt.append([kind, x["account_name"], bal, SEND if bal != 0 else HOLD])
    write_table(debt, ["Kind", "Account", "Balance $", "Draft rec"], all_debt, start_row=4)
    # summary block
    debt["F4"] = "ROLLUP"
    debt["F4"].font = SECTION_FONT
    debt["F5"] = "Credit cards"
    debt["G5"] = money(mirror["cc_total"])
    debt["F6"] = "Notes payable"
    debt["G6"] = money(mirror["np_total"])
    debt["F7"] = "Short-term financing"
    debt["G7"] = money(mirror["st_total"])
    debt["F8"] = "Total debt-like (sum)"
    debt["G8"] = money(mirror["cc_total"] + mirror["np_total"] + mirror["st_total"])
    debt["F8"].font = Font(bold=True)

    # ── 09 Mock Accrual Policy ────────────────────────────────
    pol = wb.create_sheet("09_Mock_Accrual_Policy")
    pol["A1"] = "Mock accrual policy scaffold — DECIDE with CPA (not calculated)"
    pol["A1"].font = TITLE_FONT
    pol["A2"] = "Purpose: bridge cash books → management/bank view until CPA flips company to accrual Q1 2027"
    pol["A2"].fill = NEEDS_FILL
    write_table(
        pol,
        ["Policy area", "Cash books today", "Mock option A", "Mock option B", "CPA pick", "Bank Aug 5?"],
        [
            ["Revenue", "When cash collected", "Milestone: deposit / work complete / final", "% complete by stage", "", ""],
            ["Customer deposits", "Cash in / income?", "Deferred revenue until earned", "Same as cash", "", ""],
            ["Unbilled WIP", "Invisible", "Schedule only (no BS plug)", "BS unbilled AR plug", "", ""],
            ["Job costs", "When paid", "When incurred (bill date)", "When installed", "", ""],
            ["ABC materials", "When paid on register", "Match invoice date to job", "Same as cash", "", ""],
            ["Overhead", "Period cash spend", "Leave unallocated", "Allocate by location revenue", "", ""],
            ["Credit memos", "Register/CM events", "Separate line always", "Net silently (NO)", "", ""],
            ["Cutoff", "Bank statement date", "Month-end + 3 bus days", "Calendar month only", "", ""],
        ],
        start_row=4,
    )
    pol["A16"] = "Shelf-life rule: every mock number carries policy_version. Q1 2027: swap engine to true accrual QBO; keep tab names."
    pol["A17"] = "If CPA says bank only wants cash for Aug 5 → mark 10/11 NEVER or HOLD and ship 03/06/08 + QBO P&L/BS."

    # ── 10 Mock P&L scaffold ──────────────────────────────────
    mpl = wb.create_sheet("10_Mock_P&L_Scaffold")
    mpl["A1"] = "Mock accrual P&L — EMPTY SCAFFOLD"
    mpl["A1"].font = TITLE_FONT
    mpl["A2"] = "HOLD until policy picked. Do not invent earnings for CPA meeting."
    mpl["A2"].fill = HOLD_FILL
    write_table(
        mpl,
        ["Line", "Cash basis (TBD period)", "Mock adjust", "Mock accrual", "Notes"],
        [
            ["Revenue — collected", "", "", "", "From QBO cash P&L"],
            ["+ Deferred deposit → earned", "", "", "", "Policy"],
            ["+ Unbilled WIP revenue", "", "", "", "Policy"],
            ["− Revenue collected not yet earned", "", "", "", "Policy"],
            ["Net management revenue", "", "", "", ""],
            ["Job costs — cash", "", "", "", ""],
            ["+ Accrued job costs unpaid", "", "", "", ""],
            ["Contribution", "", "", "", ""],
            ["Overhead cash", "", "", "", ""],
            ["Management operating income", "", "", "", ""],
        ],
        start_row=4,
    )

    # ── 11 Bridge ─────────────────────────────────────────────
    br = wb.create_sheet("11_Bridge_Cash_to_Mock")
    br["A1"] = "Bridge: cash earnings → mock accrual"
    br["A1"].font = TITLE_FONT
    br["A2"] = "HOLD. Required if mock P&L is sent; otherwise internal only."
    br["A2"].fill = HOLD_FILL
    write_table(
        br,
        ["Bridge step", "Amount", "Evidence", "CPA ok?"],
        [
            ["Cash net income (QBO)", "", "Cash P&L", ""],
            ["+ Deposits received not earned", "", "Deposit schedule", ""],
            ["− Deposits earned this period (prior cash)", "", "Stage/revenue policy", ""],
            ["+ Costs incurred not paid", "", "AP / ABC unmatched", ""],
            ["− Costs paid for prior period work", "", "Job cost timing", ""],
            ["= Mock accrual net", "", "", ""],
        ],
        start_row=4,
    )

    # ── 12 Data quality ───────────────────────────────────────
    dq = wb.create_sheet("12_Data_Quality_Flags")
    dq["A1"] = "Data quality & credibility flags"
    dq["A1"].font = TITLE_FONT
    write_table(
        dq,
        ["Flag", "Severity", "Evidence", "Bank impact", "Draft rec"],
        [
            ["QB AR vs AccuLynx open AR gap", "HIGH", f"QB {money(mirror.get('qb_ar'))} vs AccuLynx {money(wip['open_ar'])}", "Confusion / trust", NEEDS_CPA],
            ["Deposit collected dates missing", "HIGH", "WIP blank columns", "Can't stage cash gates cleanly", HOLD],
            ["Sales manager blank", "MED", "WIP column empty", "No manager rollup for bank", HOLD],
            ["Location cash CF unavailable", "HIGH", "Registers company-level", "Bank asked location CF", NEEDS_CPA],
            ["ABC register lag / unmatched invoices", "MED", "PEC-108 ~29% unmatched smoke", "Job cost understated", NEVER],
            ["Stage names provisional", "MED", "Generator ≠ docs/75 buckets", "Ops confusion", HOLD],
            ["Mirror last full backfill ~2026-07-27", "MED", "qbo_sync_runs", "Refresh before Aug 5", SEND],
            ["CenterPoint commercial incomplete in WIP v1", "MED", "docs/75 / PEC-110", "Service/commercial hole", HOLD],
            ["Exception queue jobs", "LOW", f"{len(wip['exceptions'])} flags in WIP", "Ops cleanup", NEVER],
        ],
        start_row=3,
    )

    # ── 13 Friday agenda ──────────────────────────────────────
    fri = wb.create_sheet("13_Friday_AR_WIP_Agenda")
    fri["A1"] = "Friday AR/WIP meeting agenda (INTERNAL — NEVER to bank)"
    fri["A1"].font = TITLE_FONT
    fri["A2"] = "Uses same spine as bank pack; different audience."
    fri["A2"].fill = NEVER_FILL
    write_table(
        fri,
        ["#", "Block (mins)", "Owner", "Artifact", "Outcome"],
        [
            ["1", "5 — Cash snapshot", "Chris/Lucinda", "03_Cash_Position + WIP Cash sheet", "Liquidity headroom"],
            ["2", "10 — AR focus", "Lucinda", "Jobs sorted by Open AR; QB aging", "Collection calls list"],
            ["3", "15 — Stage gates", "Ops", "By Stage + deposit gaps", "Which jobs stuck pre-deposit / post-complete"],
            ["4", "10 — Location heat", "Managers", "By Location", "Where cash is trapped"],
            ["5", "10 — Exceptions", "Alex/Lucinda", "Exception Queue + ABC gap", "Data fixes this week"],
            ["6", "5 — Bank pack status", "Chris", "01_Decision_Matrix post-CPA", "What ships Aug 5"],
        ],
        start_row=4,
    )
    fri["A14"] = f"Primary file for Friday: {Path(wip['path']).name}"
    fri["A15"] = "Review tomorrow: red-pen stages, AR truth, location names, anything embarrassing before Friday."

    # ── 14 CPA Meeting Script ─────────────────────────────────
    script = wb.create_sheet("14_CPA_Meeting_Script")
    script["A1"] = "30–45 min CPA meeting script (tomorrow morning)"
    script["A1"].font = TITLE_FONT
    lines = [
        "0:00 — Frame: cash books today; bank wants full CF org+location by Aug 5; accrual flip Q1 2027.",
        "0:03 — Show Cover headline facts. Pause on AR gap. Ask: how do you want that narrated?",
        "0:10 — Walk 01_Decision_Matrix top to bottom. Fill column F live.",
        "0:20 — 03 Cash + 08 Debt: confirm SEND list; kill zero-noise accounts.",
        "0:25 — 04/05 CF templates: lock period + location strategy for Aug 5 (even if location = HOLD).",
        "0:32 — 09 Mock accrual: pick revenue/cost policy OR explicitly defer mock to after Aug 5.",
        "0:40 — Capture answers on 02_Open_Questions. Photo/scan sign-off block on Cover.",
        "0:45 — End state: list of tabs that will be built for bank; owners; Chris reviews Thursday.",
    ]
    for i, line in enumerate(lines, start=3):
        script[f"A{i}"] = line
    script.column_dimensions["A"].width = 110

    out = OUT_DIR / f"cpa-bank-decision-pack-{stamp}.xlsx"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main() -> int:
    load_env()
    base = require_env("SUPABASE_URL", "PUBLIC_SUPABASE_URL").rstrip("/")
    key = require_env("SUPABASE_SERVICE_ROLE_KEY", "SUPABASE_SECRET_KEY")
    if "rnhmvcpsvtqjlffpsayu" not in base:
        raise SystemExit("Refusing non-PE Supabase host")

    wip_path = latest_wip_path()
    print("WIP pack:", wip_path)
    wip = load_wip_summary(wip_path)
    print(f"  jobs={wip['job_count']} open_ar={wip['open_ar']:,.2f} cash={wip.get('cash_total')}")

    print("loading QBO mirror registers/accounts…")
    mirror = fetch_mirror(base, key)
    print(
        f"  banks={mirror['bank_total']:,.2f} cc={mirror['cc_total']:,.2f} "
        f"np={mirror['np_total']:,.2f} qb_ar={mirror.get('qb_ar')}"
    )

    out = build_workbook(wip, mirror)
    print(f"wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
