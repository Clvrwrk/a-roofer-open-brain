#!/usr/bin/env python3
"""
Chandler WIP v1 pack (PEC-99).

Pulls AccuLynx jobs from Supabase crm_pipeline (+ joins) and a read-only QuickBooks
cash snapshot, then writes outputs/chandler-wip-v1/chandler-wip-YYYY-MM-DD.xlsx.

Hard rules:
- Never mutate QuickBooks (uses integrations/bridges/quickbooks/read-only-client.mjs).
- Do not print secrets or raw PII dumps.
- Operating stages are PROVISIONAL until Lucinda locks rules (v1.1).
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl required: pip install openpyxl") from exc

REPO = Path(__file__).resolve().parents[1]
OUT_DIR = REPO / "outputs" / "chandler-wip-v1"
PAGE = 1000
EXCLUDE_MILESTONES = {"cancelled", "dead", "unknown"}

# Provisional operating-stage overlay (exclusive). Lucinda lock = v1.1.
STAGE_ORDER = [
    "Lead",
    "Prospect",
    "Approved",
    "Approved – deposit collected",
    "Approved – work scheduled",
    "Approved – job finished",
    "WIP – supplement / RCV",
    "WIP – revenue collected, expenses open",
    "Invoiced – ready for final commission",
    "Closed",
]


def _load_env_file(path: Path, *, override: bool = False) -> None:
    if not path.exists():
        return
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        v = v.strip().strip('"').strip("'")
        if not k:
            continue
        if override or k not in os.environ:
            os.environ[k] = v


def load_env() -> None:
    """QB keys from Cleverwork master.env; PE Supabase from repo .env (overrides)."""
    _load_env_file(Path.home() / ".config/cleverwork/master.env", override=False)
    _load_env_file(REPO / ".env", override=True)
    _load_env_file(REPO / "config" / ".env", override=True)


def require_env(*names: str) -> str:
    for name in names:
        val = os.environ.get(name)
        if val:
            return val
    raise SystemExit(f"Missing env: one of {names}")


def sb_headers(key: str) -> Dict[str, str]:
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Accept": "application/json",
        "Prefer": "count=exact",
    }


def sb_get_all(base: str, key: str, table: str, select: str, filters: Optional[Dict[str, str]] = None) -> List[Dict[str, Any]]:
    """Paginate PostgREST (silent 1000-row cap)."""
    rows: List[Dict[str, Any]] = []
    start = 0
    while True:
        params = {"select": select}
        if filters:
            params.update(filters)
        qs = urllib.parse.urlencode(params, safe="(),.*:")
        url = f"{base.rstrip('/')}/rest/v1/{table}?{qs}"
        req = urllib.request.Request(
            url,
            headers={
                **sb_headers(key),
                "Range": f"{start}-{start + PAGE - 1}",
                "Range-Unit": "items",
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                chunk = json.loads(resp.read().decode())
                content_range = resp.headers.get("Content-Range", "")
        except urllib.error.HTTPError as e:
            body = e.read().decode()[:300]
            raise SystemExit(f"Supabase GET {table} failed HTTP {e.code}: {body}") from e

        if not isinstance(chunk, list):
            raise SystemExit(f"Unexpected response for {table}")
        rows.extend(chunk)
        if len(chunk) < PAGE:
            break
        start += PAGE
        # safety
        if start > 200_000:
            break
        _ = content_range
    return rows


def num(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def dedupe_pipeline(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Prefer api_sync when duplicate acculynx_job_id (mirrors exec dashboard)."""
    by_id: Dict[str, Dict[str, Any]] = {}
    passthrough: List[Dict[str, Any]] = []
    for row in rows:
        jid = row.get("acculynx_job_id")
        if not jid:
            passthrough.append(row)
            continue
        prev = by_id.get(jid)
        if prev is None:
            by_id[jid] = row
            continue
        # prefer api_sync
        if prev.get("data_source") != "api_sync" and row.get("data_source") == "api_sync":
            by_id[jid] = row
    return list(by_id.values()) + passthrough


def operating_stage(milestone: str, balance_due: float, has_insurance: bool) -> str:
    m = (milestone or "").strip().lower()
    if m in ("unassigned_lead", "assigned_lead", "lead"):
        return "Lead"
    if m == "prospect":
        return "Prospect"
    if m == "approved":
        # No deposit/schedule signals in v1 → stay at Approved
        if has_insurance and balance_due > 0:
            return "WIP – supplement / RCV"
        return "Approved"
    if m == "completed":
        if has_insurance and balance_due > 0:
            return "WIP – supplement / RCV"
        return "Approved – job finished"
    if m == "invoiced":
        if balance_due > 0 and has_insurance:
            return "WIP – supplement / RCV"
        if balance_due <= 0:
            return "Invoiced – ready for final commission"
        return "Invoiced – ready for final commission"
    if m == "closed":
        return "Closed"
    return "Prospect"  # unknown open → soft bucket


def history_dates(history_rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, str]]:
    """job_id -> {Lead, Prospect, Approved, ... ISO date} earliest per Title Case milestone."""
    out: Dict[str, Dict[str, str]] = defaultdict(dict)
    for h in history_rows:
        jid = h.get("job_id")
        name = (h.get("milestone_name") or "").strip()
        dt = h.get("milestone_date")
        if not jid or not name or not dt:
            continue
        key = name.title() if name.lower() != name else name
        # normalize common
        canon = {
            "lead": "Lead",
            "prospect": "Prospect",
            "approved": "Approved",
            "completed": "Completed",
            "invoiced": "Invoiced",
            "closed": "Closed",
            "cancelled": "Cancelled",
        }.get(name.lower(), name)
        iso = str(dt)[:10]
        prev = out[jid].get(canon)
        if prev is None or iso < prev:
            out[jid][canon] = iso
    return out


def open_ar_by_job(invoice_rows: List[Dict[str, Any]]) -> Dict[str, float]:
    ar: Dict[str, float] = defaultdict(float)
    for inv in invoice_rows:
        jid = inv.get("job_id")
        if not jid:
            continue
        bal = num(inv.get("balance_due"))
        if bal > 0:
            ar[jid] += bal
    return ar


def qb_cash_snapshot() -> Dict[str, Any]:
    """Run node read-only helper for company + BalanceSheet cash-ish totals."""
    script = REPO / "integrations/bridges/quickbooks/read-only-client.mjs"
    # Inline node snippet importing the client
    node = r"""
import { refreshAccessToken, qboGet, realmId, apiBase, isReadOnlyMode } from './integrations/bridges/quickbooks/read-only-client.mjs';
const realm = realmId();
const { accessToken } = await refreshAccessToken();
const info = await qboGet(accessToken, `companyinfo/${realm}`);
const company = info.CompanyInfo || {};
let balanceSheet = null;
try {
  balanceSheet = await qboGet(accessToken, 'reports/BalanceSheet');
} catch (e) {
  balanceSheet = { error: String(e.message || e) };
}
let arAging = null;
try {
  arAging = await qboGet(accessToken, 'reports/AgedReceivableDetail');
} catch (e) {
  try {
    arAging = await qboGet(accessToken, 'reports/AgedReceivables');
  } catch (e2) {
    arAging = { error: String(e2.message || e2) };
  }
}
function sumBankish(report) {
  // Best-effort walk of BalanceSheet Rows for account names containing Cash/Bank/Checking
  let total = null;
  const hits = [];
  function walk(node) {
    if (!node) return;
    if (Array.isArray(node)) { node.forEach(walk); return; }
    if (typeof node !== 'object') return;
    const header = node.Header?.ColData || node.ColData;
    const summary = node.Summary?.ColData;
    const name = header?.[0]?.value || '';
    const amtStr = summary?.[1]?.value ?? header?.[1]?.value;
    if (name && /cash|bank|checking|money market|savings/i.test(name) && amtStr && amtStr !== '') {
      const n = Number(String(amtStr).replace(/[,$]/g,''));
      if (!Number.isNaN(n)) hits.push({ name, amount: n });
    }
    if (node.Rows?.Row) walk(node.Rows.Row);
    if (node.Row) walk(node.Row);
  }
  walk(report?.Rows?.Row || report);
  if (hits.length) total = hits.reduce((s, h) => s + h.amount, 0);
  return { total, hits: hits.slice(0, 40) };
}
const cash = sumBankish(balanceSheet);
console.log(JSON.stringify({
  readOnly: isReadOnlyMode(),
  apiBase: apiBase(),
  realmId: realm,
  companyName: company.CompanyName || null,
  legalName: company.LegalName || null,
  country: company.Country || null,
  cashTotal: cash.total,
  cashAccounts: cash.hits,
  balanceSheetError: balanceSheet?.error || null,
  arAgingError: arAging?.error || null,
}));
"""
    env = os.environ.copy()
    proc = subprocess.run(
        ["node", "--input-type=module", "-e", node],
        cwd=str(REPO),
        env=env,
        capture_output=True,
        text=True,
        timeout=120,
    )
    if proc.returncode != 0:
        return {
            "error": (proc.stderr or proc.stdout or "qb snapshot failed")[:400],
            "readOnly": True,
        }
    # last JSON line
    lines = [ln for ln in proc.stdout.splitlines() if ln.strip().startswith("{")]
    if not lines:
        return {"error": "no JSON from qb snapshot", "raw_tail": proc.stdout[-200:]}
    return json.loads(lines[-1])


def autosize(ws, max_width: int = 42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col[:200]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def write_sheet(ws, headers: List[str], rows: Iterable[List[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws)
    autosize(ws)


def main() -> int:
    load_env()
    base = require_env("SUPABASE_URL", "PUBLIC_SUPABASE_URL").rstrip("/")
    key = require_env("SUPABASE_SERVICE_ROLE_KEY", "SUPABASE_SECRET_KEY")
    if "rnhmvcpsvtqjlffpsayu" not in base:
        raise SystemExit(
            f"Refusing non-PE Supabase host ({base.split('//')[-1]}). "
            "Expected rnhmvcpsvtqjlffpsayu — check repo .env"
        )

    print("loading crm_pipeline…")
    pipeline = sb_get_all(
        base,
        key,
        "crm_pipeline",
        select=(
            "id,acculynx_job_id,job_name,client_name,client_job_number,"
            "location_city,location_state,location_zip,current_milestone,"
            "primary_salesperson,contract_amount,primary_estimate_amount,balance_due,"
            "lead_date,approved_date,milestone_date,created_at,updated_at,"
            "insurance_company,insurance_claim_number,insurance_claim_filed,"
            "job_category,data_source"
        ),
    )
    print(f"  rows={len(pipeline)}")
    pipeline = dedupe_pipeline(pipeline)
    print(f"  after_dedupe={len(pipeline)}")

    print("loading acculynx_jobs account_key…")
    jobs = sb_get_all(base, key, "acculynx_jobs", select="id,account_key,job_category_name,current_milestone")
    job_map = {j["id"]: j for j in jobs if j.get("id")}
    print(f"  jobs={len(job_map)}")

    print("loading milestone history…")
    history = sb_get_all(
        base,
        key,
        "acculynx_job_milestone_history",
        select="job_id,milestone_name,milestone_date",
    )
    hist = history_dates(history)
    print(f"  history_rows={len(history)} jobs_with_dates={len(hist)}")

    print("loading open invoice AR…")
    invoices = sb_get_all(
        base,
        key,
        "acculynx_invoices",
        select="job_id,balance_due,total_price,current_invoice_state",
    )
    ar_map = open_ar_by_job(invoices)
    print(f"  invoices={len(invoices)} jobs_with_open_ar={sum(1 for v in ar_map.values() if v > 0)}")

    print("loading QB cash snapshot (read-only)…")
    qb = qb_cash_snapshot()
    print(
        "  qb=",
        qb.get("companyName"),
        "cashTotal=",
        qb.get("cashTotal"),
        "readOnly=",
        qb.get("readOnly"),
        "err=",
        qb.get("error") or qb.get("balanceSheetError"),
    )

    # Build job rows
    job_rows: List[List[Any]] = []
    stage_agg: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"count": 0, "contract": 0.0, "ar": 0.0})
    loc_agg: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"count": 0, "contract": 0.0, "ar": 0.0})
    rep_agg: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"count": 0, "contract": 0.0, "ar": 0.0})
    exceptions: List[List[Any]] = []

    for row in pipeline:
        milestone = (row.get("current_milestone") or "").strip().lower()
        if milestone in EXCLUDE_MILESTONES:
            continue
        jid = row.get("acculynx_job_id")
        aj = job_map.get(jid or "", {})
        account = aj.get("account_key") or "unknown"
        contract = num(row.get("contract_amount"))
        estimate = num(row.get("primary_estimate_amount"))
        pipe_bal = num(row.get("balance_due"))
        ar = ar_map.get(jid or "", 0.0)
        # Prefer invoice AR when present
        bal = ar if ar > 0 else pipe_bal
        has_ins = bool(row.get("insurance_company") or row.get("insurance_claim_number") or row.get("insurance_claim_filed"))
        stage = operating_stage(milestone, bal, has_ins)
        dates = hist.get(jid or "", {})
        value_for_stage = contract if contract > 0 else estimate

        job_rows.append(
            [
                account,
                row.get("client_job_number"),
                row.get("client_name") or row.get("job_name"),
                row.get("job_name"),
                milestone,
                stage,
                "provisional — Lucinda lock v1.1",
                row.get("primary_salesperson") or "(unassigned)",
                "",  # sales_manager TBD
                row.get("job_category") or aj.get("job_category_name"),
                contract,
                estimate,
                bal,
                dates.get("Lead") or row.get("lead_date"),
                dates.get("Prospect"),
                dates.get("Approved") or row.get("approved_date"),
                "",  # deposit collected — data gap
                "",  # work scheduled — data gap
                dates.get("Completed"),
                dates.get("Invoiced"),
                dates.get("Closed"),
                row.get("location_city"),
                row.get("location_state"),
                row.get("location_zip"),
                row.get("insurance_company"),
                row.get("insurance_claim_number"),
                f"https://my.acculynx.com/jobs/{jid}" if jid else "",
                row.get("data_source"),
                jid,
            ]
        )

        stage_agg[stage]["count"] += 1
        stage_agg[stage]["contract"] += value_for_stage
        stage_agg[stage]["ar"] += bal
        loc_agg[account]["count"] += 1
        loc_agg[account]["contract"] += value_for_stage
        loc_agg[account]["ar"] += bal
        rep = row.get("primary_salesperson") or "(unassigned)"
        rep_agg[rep]["count"] += 1
        rep_agg[rep]["contract"] += value_for_stage
        rep_agg[rep]["ar"] += bal

        # Exceptions
        if stage == "WIP – supplement / RCV":
            exceptions.append(["WIP supplement/RCV candidate", account, rep, row.get("job_name"), bal, contract, jid])
        if milestone in ("approved", "completed", "invoiced") and contract <= 0:
            exceptions.append(["Approved+ with $0 contract", account, rep, row.get("job_name"), bal, contract, jid])
        if not row.get("primary_salesperson") and milestone not in ("unassigned_lead",):
            exceptions.append(["Missing salesperson", account, "", row.get("job_name"), bal, contract, jid])

    # Sort jobs: stage order then location then name
    stage_rank = {s: i for i, s in enumerate(STAGE_ORDER)}
    job_rows.sort(key=lambda r: (stage_rank.get(r[5], 99), str(r[0]), str(r[2] or "")))

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = date.today().isoformat()
    out_path = OUT_DIR / f"chandler-wip-{stamp}.xlsx"

    wb = Workbook()

    # Cover
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Chandler WIP Pack — v1 (PROVISIONAL)"
    cover["A1"].font = Font(bold=True, size=16)
    cover["A3"] = "Generated"
    cover["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    cover["A4"] = "Linear"
    cover["B4"] = "PEC-99"
    cover["A5"] = "QB access"
    cover["B5"] = "read_only / mirror-only (PEC-98)"
    cover["A6"] = "QB company"
    cover["B6"] = qb.get("companyName") or qb.get("error") or "(unavailable)"
    cover["A7"] = "QB realm"
    cover["B7"] = qb.get("realmId") or ""
    cover["A8"] = "Cash-like total (BalanceSheet walk)"
    cover["B8"] = qb.get("cashTotal")
    cover["A10"] = "Notes"
    cover["A11"] = (
        "Operating stages are best-effort from AccuLynx milestones + insurance/AR signals. "
        "Deposit collected / work scheduled columns are blank (data not in sync yet). "
        "Sales manager blank until Lucinda map. Cancelled/dead excluded. "
        "One row per job (api_sync preferred). Values counted once in current stage only."
    )
    cover["A11"].alignment = Alignment(wrap_text=True)
    cover.merge_cells("A11:F14")
    cover.column_dimensions["A"].width = 36
    cover.column_dimensions["B"].width = 48

    # Jobs
    jobs_ws = wb.create_sheet("Jobs")
    write_sheet(
        jobs_ws,
        [
            "Location (account_key)",
            "Job #",
            "Client / display",
            "Job name",
            "AccuLynx milestone",
            "Operating stage (v1)",
            "Stage confidence",
            "Salesperson",
            "Sales manager",
            "Job category",
            "Contract $",
            "Estimate $",
            "Open AR $",
            "Lead date",
            "Prospect date",
            "Approved date",
            "Deposit collected",
            "Work scheduled",
            "Job finished",
            "Invoiced date",
            "Closed date",
            "City",
            "State",
            "Zip",
            "Insurance co",
            "Claim #",
            "AccuLynx URL",
            "Data source",
            "AccuLynx job id",
        ],
        job_rows,
    )

    # By Stage
    stage_ws = wb.create_sheet("By Stage")
    stage_rows = []
    for s in STAGE_ORDER:
        a = stage_agg.get(s) or {"count": 0, "contract": 0.0, "ar": 0.0}
        stage_rows.append([s, a["count"], round(a["contract"], 2), round(a["ar"], 2)])
    # any unexpected stages
    for s, a in sorted(stage_agg.items()):
        if s not in stage_rank:
            stage_rows.append([s, a["count"], round(a["contract"], 2), round(a["ar"], 2)])
    write_sheet(stage_ws, ["Operating stage", "Jobs", "Contract/Estimate $", "Open AR $"], stage_rows)

    # By Location
    loc_ws = wb.create_sheet("By Location")
    loc_rows = [
        [k, v["count"], round(v["contract"], 2), round(v["ar"], 2)]
        for k, v in sorted(loc_agg.items(), key=lambda kv: -kv[1]["contract"])
    ]
    write_sheet(loc_ws, ["Location", "Jobs", "Contract/Estimate $", "Open AR $"], loc_rows)

    # By Salesperson
    rep_ws = wb.create_sheet("By Salesperson")
    rep_rows = [
        [k, v["count"], round(v["contract"], 2), round(v["ar"], 2)]
        for k, v in sorted(rep_agg.items(), key=lambda kv: -kv[1]["contract"])
    ]
    write_sheet(rep_ws, ["Salesperson", "Jobs", "Contract/Estimate $", "Open AR $"], rep_rows)

    # Cash / Runway
    cash_ws = wb.create_sheet("Cash Snapshot (QB)")
    cash_ws["A1"] = "QuickBooks cash snapshot (read-only)"
    cash_ws["A1"].font = Font(bold=True, size=14)
    cash_ws["A3"] = "Company"
    cash_ws["B3"] = qb.get("companyName")
    cash_ws["A4"] = "Legal name"
    cash_ws["B4"] = qb.get("legalName")
    cash_ws["A5"] = "Realm"
    cash_ws["B5"] = qb.get("realmId")
    cash_ws["A6"] = "API base"
    cash_ws["B6"] = qb.get("apiBase")
    cash_ws["A7"] = "Read-only enforced"
    cash_ws["B7"] = qb.get("readOnly")
    cash_ws["A8"] = "Cash-like total"
    cash_ws["B8"] = qb.get("cashTotal")
    cash_ws["A9"] = "Notes"
    cash_ws["B9"] = (
        "Cash-like total is a best-effort walk of BalanceSheet rows matching Cash/Bank/Checking. "
        "Burn / runway months need payroll+opex mapping (v1.1 with Lucinda). "
        "This sheet never writes to QB."
    )
    cash_ws["B9"].alignment = Alignment(wrap_text=True)
    cash_ws["A11"] = "Cash/Bank account lines (sample)"
    cash_ws.append([])
    cash_ws.append(["Account", "Amount"])
    for hit in qb.get("cashAccounts") or []:
        cash_ws.append([hit.get("name"), hit.get("amount")])
    if qb.get("error") or qb.get("balanceSheetError"):
        cash_ws.append([])
        cash_ws.append(["Errors", qb.get("error") or qb.get("balanceSheetError")])
    autosize(cash_ws)

    # Exceptions
    ex_ws = wb.create_sheet("Exception Queue")
    write_sheet(
        ex_ws,
        ["Flag", "Location", "Salesperson", "Job", "Open AR $", "Contract $", "Job id"],
        exceptions[:5000],
    )

    # Data gaps
    gaps = wb.create_sheet("Data Gaps (v1)")
    write_sheet(
        gaps,
        ["Gap", "Impact", "v1 workaround", "Owner"],
        [
            ["Sales manager", "No manager rollup", "Blank column", "Lucinda / Ops"],
            ["Deposit collected date/amount", "Can't split Approved deposit bucket", "Blank cols; stage stays Approved", "AccuLynx + Lucinda"],
            ["Work scheduled date", "Can't split Approved scheduled bucket", "Blank cols", "Ops schedule source"],
            ["WIP expenses-open after cash collected", "Needs QB Class↔job map + AP", "Not auto-staged in v1", "Accounting + PEC-69"],
            ["Centerpoint commercial", "Service/commercial WIP incomplete", "Out of scope v1", "PE-CC"],
            ["crm_pipeline.balance_due unreliable", "AR may be $0 on pipeline", "Prefer acculynx_invoices open balance", "Engineering"],
        ],
    )

    wb.save(out_path)
    print(f"wrote {out_path}")
    print(f"open_jobs={len(job_rows)} exceptions={len(exceptions)}")
    print("stage_counts=", {s: stage_agg[s]["count"] for s in STAGE_ORDER if stage_agg[s]["count"]})
    return 0


if __name__ == "__main__":
    sys.exit(main())
